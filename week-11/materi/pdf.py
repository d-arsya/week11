import PyPDF2
import pdfplumber

print("PyPDF2 dan pdfplumber berhasil diimport!")

print("=== READING PDF WITH PYPDF2 ===")

# Note: We'll create a simple text-based example
# In real scenario, you would have actual PDF files

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

def create_sample_pdf():
    c = canvas.Canvas("hasil_lab.pdf", pagesize=letter)
    
    # Title
    c.setFont("Helvetica-Bold", 16)
    c.drawString(100, 750, "HASIL PEMERIKSAAN LABORATORIUM")
    
    # Patient info
    c.setFont("Helvetica", 12)
    c.drawString(100, 700, "Nama Pasien: Ahmad Rizky Pratama")
    c.drawString(100, 680, "No. RM: RM-2024-001")
    c.drawString(100, 660, "Tanggal: 16 Januari 2024")
    
    # Lab results
    c.drawString(100, 620, "HASIL PEMERIKSAAN:")
    c.drawString(120, 600, "Glukosa Puasa: 180 mg/dL (Tinggi)")
    c.drawString(120, 580, "Kolesterol Total: 250 mg/dL (Tinggi)")
    c.drawString(120, 560, "Trigliserida: 200 mg/dL (Normal)")
    c.drawString(120, 540, "HDL: 45 mg/dL (Normal)")
    c.drawString(120, 520, "LDL: 160 mg/dL (Tinggi)")
    
    c.save()
    print("Sample PDF created: hasil_lab.pdf")

create_sample_pdf()

# Read PDF with PyPDF2
print("\n=== READING PDF ===")
with open("hasil_lab.pdf", "rb") as file:
    pdf_reader = PyPDF2.PdfReader(file)
    
    print(f"Number of pages: {len(pdf_reader.pages)}")
    
    # Extract text from first page
    page = pdf_reader.pages[0]
    text = page.extract_text()
    
    print("\nExtracted text:")
    print(text)

print("\n=== READING PDF WITH PDFPLUMBER ===")

# pdfplumber is better for extracting tables
with pdfplumber.open("hasil_lab.pdf") as pdf:
    first_page = pdf.pages[0]
    
    # Extract text
    text = first_page.extract_text()
    print("Extracted text:")
    print(text)
    
    # Extract tables (if any)
    tables = first_page.extract_tables()
    if tables:
        print("\nTables found:")
        for table in tables:
            print(table)
    else:
        print("\nNo tables found in this PDF")

print("\n=== PARSING LAB RESULTS ===")

def parse_lab_results(pdf_path):
    """Extract lab results from PDF"""
    results = {}
    
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text()
        
        # Parse patient info
        for line in text.split('\n'):
            if "Nama Pasien:" in line:
                results['nama'] = line.split(":")[1].strip()
            elif "No. RM:" in line:
                results['no_rm'] = line.split(":")[1].strip()
            elif "Tanggal:" in line:
                results['tanggal'] = line.split(":")[1].strip()
            elif "Glukosa" in line:
                # Extract numeric value
                parts = line.split(":")
                if len(parts) > 1:
                    value_part = parts[1].strip().split()[0]
                    results['glukosa'] = value_part
            elif "Kolesterol Total" in line:
                parts = line.split(":")
                if len(parts) > 1:
                    value_part = parts[1].strip().split()[0]
                    results['kolesterol'] = value_part
    
    return results

# Parse the PDF
lab_data = parse_lab_results("hasil_lab.pdf")
print("\nParsed lab results:")
for key, value in lab_data.items():
    print(f"{key}: {value}")

print("\n=== PDF TO EXCEL CONVERSION ===")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def pdf_to_excel(pdf_path, excel_path):
    """Convert lab results from PDF to Excel"""
    
    # Parse PDF
    lab_data = parse_lab_results(pdf_path)
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil Lab"
    
    # Headers
    ws.append(["Field", "Value"])
    
    # Data
    for key, value in lab_data.items():
        ws.append([key.title(), value])
    
    # Format
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    wb.save(excel_path)
    print(f"Excel file created: {excel_path}")

# Convert
pdf_to_excel("hasil_lab.pdf", "hasil_lab_converted.xlsx")

print("\n=== BATCH PDF PROCESSING ===")

import os
from openpyxl import Workbook

def batch_process_pdfs(pdf_folder, output_excel):
    """Process multiple PDF files and consolidate to Excel"""

    wb = Workbook()
    ws = wb.active
    ws.title = "All Lab Results"

    # Headers
    ws.append(["No. RM", "Nama", "Tanggal", "Glukosa", "Kolesterol"])

    # Process each PDF
    pdf_files = [f for f in os.listdir(pdf_folder) if f.endswith('.pdf')]

    for pdf_file in pdf_files:
        pdf_path = os.path.join(pdf_folder, pdf_file)

        try:
            lab_data = parse_lab_results(pdf_path)
            ws.append([
                lab_data.get('no_rm', ''),
                lab_data.get('nama', ''),
                lab_data.get('tanggal', ''),
                lab_data.get('glukosa', ''),
                lab_data.get('kolesterol', '')
            ])
            print(f"Processed: {pdf_file}")
        except Exception as e:
            print(f"Error processing {pdf_file}: {e}")

    # Format headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    wb.save(output_excel)
    print(f"\nAll results saved to: {output_excel}")
    print(f"Total files processed: {len(pdf_files)}")

# Example usage (assuming you have multiple PDFs in a folder)
# batch_process_pdfs("lab_results", "all_lab_results.xlsx")
print("Batch processing function ready to use!")
