import pdfplumber
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from docx import Document
from docx.shared import Inches, Pt, RGBColor
from docx.enum.text import WD_ALIGN_PARAGRAPH
from openpyxl import load_workbook
from docx import Document
from datetime import datetime
print("\n=== ERROR HANDLING ===")

def safe_read_excel(file_path):
    """Safely read Excel with error handling"""
    try:
        wb = load_workbook(file_path)
        print(f"✓ Successfully loaded: {file_path}")
        return wb
    except FileNotFoundError:
        print(f"✗ Error: File {file_path} not found")
        return None
    except PermissionError:
        print(f"✗ Error: No permission to access {file_path}")
        return None
    except Exception as e:
        print(f"✗ Error loading file: {e}")
        return None

def safe_read_pdf(file_path):
    """Safely read PDF with error handling"""
    try:
        with pdfplumber.open(file_path) as pdf:
            print(f"✓ Successfully loaded PDF: {file_path}")
            return pdf.pages[0].extract_text()
    except FileNotFoundError:
        print(f"✗ Error: PDF {file_path} not found")
        return None
    except Exception as e:
        print(f"✗ Error reading PDF: {e}")
        return None

# Test error handling
wb = safe_read_excel("data_pasien.xlsx")
wb_missing = safe_read_excel("file_tidak_ada.xlsx")

text = safe_read_pdf("hasil_lab.pdf")
text_missing = safe_read_pdf("pdf_tidak_ada.pdf")

print("\n=== DATA VALIDATION ===")

def validate_patient_data(data):
    """Validate patient data before processing"""
    errors = []

    # Check required fields
    required_fields = ['no_rm', 'nama', 'umur']
    for field in required_fields:
        if not data.get(field):
            errors.append(f"Missing required field: {field}")

    # Validate age
    try:
        umur = int(data.get('umur', 0))
        if umur <= 0 or umur > 150:
            errors.append("Invalid age: must be between 1-150")
    except ValueError:
        errors.append("Age must be a number")

    # Validate No. RM format
    no_rm = data.get('no_rm', '')
    if not no_rm.startswith('RM-'):
        errors.append("Invalid No. RM format (should start with RM-)")

    return errors

# Test validation
test_data = [
    {'no_rm': 'RM-001', 'nama': 'Ahmad', 'umur': 28},
    {'no_rm': '', 'nama': 'Sari', 'umur': 35},
    {'no_rm': 'RM-003', 'nama': 'Budi', 'umur': 200},
    {'no_rm': 'ABC-004', 'nama': 'Nina', 'umur': 29}
]

for i, data in enumerate(test_data, 1):
    print(f"\nValidating data {i}: {data}")
    errors = validate_patient_data(data)
    if errors:
        print("  ✗ Validation errors:")
        for error in errors:
            print(f"    - {error}")
    else:
        print("  ✓ Data valid")
        
print("\n=== LOGGING AND PROGRESS ===")

import logging
from datetime import datetime

# Setup logging
logging.basicConfig(
    filename='document_processing.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def process_with_logging(excel_path):
    """Process Excel with logging"""
    logging.info(f"Starting to process: {excel_path}")

    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        total_rows = ws.max_row - 1  # Exclude header
        logging.info(f"Found {total_rows} rows to process")

        processed = 0
        errors = 0

        for row in range(2, ws.max_row + 1):
            try:
                # Process row
                no_rm = ws.cell(row=row, column=1).value
                nama = ws.cell(row=row, column=2).value

                # Simulate processing
                print(f"Processing: {no_rm} - {nama}")
                processed += 1

                logging.info(f"Processed: {no_rm} - {nama}")

            except Exception as e:
                errors += 1
                logging.error(f"Error processing row {row}: {e}")

        logging.info(f"Processing complete: {processed} success, {errors} errors")
        print(f"\n✓ Processed: {processed} records")
        print(f"✗ Errors: {errors}")

    except Exception as e:
        logging.error(f"Fatal error: {e}")
        print(f"✗ Fatal error: {e}")

# Process with logging
process_with_logging("data_pasien.xlsx")
print("\nCheck 'document_processing.log' for detailed logs")