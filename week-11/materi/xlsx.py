# Installation
# pip install openpyxl

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

print("openpyxl berhasil diimport!")
print(f"Version: {openpyxl.__version__}")

print("=== READING EXCEL FILES ===")

# Create sample Excel file first
wb = Workbook()
ws = wb.active
ws.title = "Data Pasien"

# Headers
headers = ["No. RM", "Nama", "Umur", "Gender", "Diagnosa", "Biaya"]
ws.append(headers)

# Sample data
patients = [
    ["RM-001", "Ahmad Rizky", 28, "L", "Diabetes Mellitus", 750000],
    ["RM-002", "Sari Dewi", 35, "P", "Hipertensi", 450000],
    ["RM-003", "Budi Santoso", 42, "L", "Gastritis", 350000],
    ["RM-004", "Nina Wulan", 29, "P", "Anemia", 500000],
    ["RM-005", "Eko Prasetyo", 38, "L", "Diabetes Mellitus", 800000]
]

for patient in patients:
    ws.append(patient)

wb.save("data_pasien.xlsx")
print("File data_pasien.xlsx berhasil dibuat")

# Now read the file
print("\n=== MEMBACA FILE EXCEL ===")
wb = load_workbook("data_pasien.xlsx")
ws = wb.active

print(f"Worksheet name: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}")
print(f"Max column: {ws.max_column}")

# Read all data
print("\nSemua data:")
for row in ws.iter_rows(min_row=1, values_only=True):
    print(row)


print("\n=== ACCESSING CELLS ===")

wb = load_workbook("data_pasien.xlsx")
ws = wb.active

# Access by cell reference
print(f"Cell A1: {ws['A1'].value}")
print(f"Cell B2: {ws['B2'].value}")

# Access by row and column
print(f"Cell (1,1): {ws.cell(row=1, column=1).value}")
print(f"Cell (2,3): {ws.cell(row=2, column=3).value}")

# Read specific row
print("\nData pasien pertama:")
for col in range(1, 7):
    cell_value = ws.cell(row=2, column=col).value
    print(f"  Column {col}: {cell_value}")

# Read specific column
print("\nSemua nama pasien:")
for row in range(2, ws.max_row + 1):
    nama = ws.cell(row=row, column=2).value
    print(f"  {nama}")


print("\n=== FILTERING DATA ===")

wb = load_workbook("data_pasien.xlsx")
ws = wb.active

# Filter pasien diabetes
print("Pasien dengan Diabetes:")
for row in range(2, ws.max_row + 1):
    diagnosa = ws.cell(row=row, column=5).value
    if "Diabetes" in diagnosa:
        no_rm = ws.cell(row=row, column=1).value
        nama = ws.cell(row=row, column=2).value
        biaya = ws.cell(row=row, column=6).value
        print(f"  {no_rm} - {nama} - Rp {biaya:,}")

# Filter by age
print("\nPasien umur > 35:")
for row in range(2, ws.max_row + 1):
    umur = ws.cell(row=row, column=3).value
    if umur > 35:
        nama = ws.cell(row=row, column=2).value
        print(f"  {nama} - {umur} tahun")

# Calculate total
print("\n=== KALKULASI ===")
total_biaya = 0
for row in range(2, ws.max_row + 1):
    biaya = ws.cell(row=row, column=6).value
    total_biaya += biaya

rata_rata = total_biaya / (ws.max_row - 1)
print(f"Total biaya: Rp {total_biaya:,}")
print(f"Rata-rata biaya: Rp {rata_rata:,.0f}")

print("\n=== WRITING TO EXCEL ===")

# Create new workbook
wb = Workbook()
ws = wb.active
ws.title = "Laporan Diabetes"

# Write headers
ws.append(["No. RM", "Nama", "Umur", "Biaya", "Kategori"])

# Read from original file and filter
wb_source = load_workbook("data_pasien.xlsx")
ws_source = wb_source.active

for row in range(2, ws_source.max_row + 1):
    diagnosa = ws_source.cell(row=row, column=5).value
    if "Diabetes" in diagnosa:
        no_rm = ws_source.cell(row=row, column=1).value
        nama = ws_source.cell(row=row, column=2).value
        umur = ws_source.cell(row=row, column=3).value
        biaya = ws_source.cell(row=row, column=6).value
        
        # Add category
        kategori = "Mahal" if biaya > 700000 else "Normal"
        
        ws.append([no_rm, nama, umur, biaya, kategori])

wb.save("laporan_diabetes.xlsx")
print("File laporan_diabetes.xlsx berhasil dibuat")

print("\n=== FORMATTING CELLS ===")

wb = load_workbook("laporan_diabetes.xlsx")
ws = wb.active

# Format header
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Format biaya column
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=4)
    cell.number_format = 'Rp #,##0'

# Auto-adjust column width
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save("laporan_diabetes_formatted.xlsx")
print("File formatted berhasil dibuat")

print("\n=== MULTIPLE SHEETS ===")

wb = Workbook()

# Create sheets for different diagnoses
ws_diabetes = wb.active
ws_diabetes.title = "Diabetes"

ws_hipertensi = wb.create_sheet("Hipertensi")
ws_lainnya = wb.create_sheet("Lainnya")

# Add headers to all sheets
headers = ["No. RM", "Nama", "Umur", "Gender", "Biaya"]
ws_diabetes.append(headers)
ws_hipertensi.append(headers)
ws_lainnya.append(headers)

# Read and distribute data
wb_source = load_workbook("data_pasien.xlsx")
ws_source = wb_source.active

for row in range(2, ws_source.max_row + 1):
    no_rm = ws_source.cell(row=row, column=1).value
    nama = ws_source.cell(row=row, column=2).value
    umur = ws_source.cell(row=row, column=3).value
    gender = ws_source.cell(row=row, column=4).value
    diagnosa = ws_source.cell(row=row, column=5).value
    biaya = ws_source.cell(row=row, column=6).value
    
    row_data = [no_rm, nama, umur, gender, biaya]
    
    if "Diabetes" in diagnosa:
        ws_diabetes.append(row_data)
    elif "Hipertensi" in diagnosa:
        ws_hipertensi.append(row_data)
    else:
        ws_lainnya.append(row_data)

wb.save("laporan_per_diagnosa.xlsx")
print("File dengan multiple sheets berhasil dibuat")
print(f"Sheet names: {wb.sheetnames}")
